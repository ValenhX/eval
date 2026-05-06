"""Module 4 : ExcelExporter

Charge les KPIs extraits dans un DataFrame pandas, calcule les statistiques
du groupe de comparables, et exporte un fichier Excel .xlsx formaté.

Pipeline :
1. Conversion de la liste de dictionnaires en DataFrame avec renommage des colonnes.
2. Calcul des statistiques de synthèse du peer group (médiane, moyenne, min, max).
3. Écriture du fichier Excel via openpyxl avec :
   - En-tête stylisée (fond coloré, texte en gras).
   - Formats numériques (pourcentage pour marge brute, monétaire pour CAPEX).
   - Lignes de synthèse en bas de tableau.
   - Largeurs de colonnes auto-ajustées.
   - Première ligne figée (freeze pane).
"""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_DEFAULT_OUTPUT_DIR = Path("data")
_DEFAULT_OUTPUT_NAME = "benchmark_kpis.xlsx"

# ---------------------------------------------------------------------------
# Constantes de mise en forme
# ---------------------------------------------------------------------------

# Correspondance colonnes internes → labels Excel
_COLUMN_MAP: dict[str, str] = {
    "nom": "Société",
    "ticker": "Ticker",
    "annee_exercice": "Exercice",
    "devise": "Devise",
    "capex": "CAPEX (M)",
    "marge_brute": "Marge Brute",
    "ratio_endettement": "Ratio Endettement",
    "kpi_source": "Source IA",
}

# Colonnes numériques et leur format openpyxl
_FORMAT_PERCENT = "0.00%"
_FORMAT_MONETARY = '#,##0.0'
_FORMAT_RATIO = "0.00"
_FORMAT_YEAR = "0"

_NUMERIC_FORMATS: dict[str, str] = {
    "CAPEX (M)": _FORMAT_MONETARY,
    "Marge Brute": _FORMAT_PERCENT,
    "Ratio Endettement": _FORMAT_RATIO,
    "Exercice": _FORMAT_YEAR,
}

# Styles visuels
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")   # bleu marine
_SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="D6DCE4")  # gris clair
_SEPARATOR_FILL = PatternFill(fill_type="solid", fgColor="BDD7EE") # bleu pâle
_INITIAL_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")   # jaune pâle
_PEERS_SEPARATOR_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")  # vert pâle
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_DATA_FONT = Font(name="Calibri", size=10)
_SUMMARY_FONT = Font(bold=True, name="Calibri", size=10)
_SEPARATOR_FONT = Font(bold=True, italic=True, name="Calibri", size=10, color="1F3864")
_INITIAL_FONT = Font(bold=True, name="Calibri", size=10, color="7F6000")
_PEERS_SEPARATOR_FONT = Font(bold=True, italic=True, name="Calibri", size=10, color="375623")

_THIN_BORDER_SIDE = Side(style="thin", color="BFBFBF")
_THIN_BORDER = Border(
    left=_THIN_BORDER_SIDE,
    right=_THIN_BORDER_SIDE,
    top=_THIN_BORDER_SIDE,
    bottom=_THIN_BORDER_SIDE,
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")

# Libellés des lignes de synthèse (ordre d'affichage)
_SUMMARY_LABELS: list[tuple[str, str]] = [
    ("Médiane", "median"),
    ("Moyenne", "mean"),
    ("Min", "min"),
    ("Max", "max"),
]

# Colonnes numériques utilisées pour les statistiques
_STAT_COLUMNS: list[str] = ["CAPEX (M)", "Marge Brute", "Ratio Endettement"]


# ---------------------------------------------------------------------------
# ExcelExporter
# ---------------------------------------------------------------------------


class ExcelExporter:
    """Module 4 — Exporte le benchmark des KPIs dans un fichier Excel formaté.

    Chaque colonne reçoit le format numérique adapté à son type de donnée.
    Des lignes de synthèse (médiane, moyenne, min, max) sont ajoutées en bas
    pour faciliter la lecture comparative du peer group.

    Example:
        >>> exporter = ExcelExporter()
        >>> path = exporter.export([
        ...     {"nom": "Apple Inc.", "ticker": "AAPL", "annee_exercice": 2023,
        ...      "devise": "USD", "capex": 11455.0, "marge_brute": 0.43,
        ...      "ratio_endettement": 0.8, "kpi_source": "Cash Flow Statement"},
        ... ])
        >>> print(path)  # data/benchmark_kpis.xlsx
    """

    def __init__(self, output_dir: Path = _DEFAULT_OUTPUT_DIR) -> None:
        """Args:
            output_dir: Répertoire de destination du fichier Excel.
        """
        self._output_dir = output_dir

    def export(
        self,
        companies: list[dict],
        filename: str = _DEFAULT_OUTPUT_NAME,
        initial_company: Optional[dict] = None,
    ) -> Path:
        """Génère le fichier Excel du benchmark.

        Args:
            companies: Sortie du Module 3 — liste de dicts contenant les KPIs.
            filename: Nom du fichier Excel de sortie.
            initial_company: Société de référence (enrichie avec KPIs) à afficher
                en tête de tableau avec un style distinct.

        Returns:
            Chemin absolu du fichier Excel généré.

        Raises:
            ValueError: Si la liste d'entreprises est vide.
            OSError: Si l'écriture du fichier échoue.
        """
        if not companies:
            raise ValueError("La liste d'entreprises est vide — rien à exporter.")

        self._output_dir.mkdir(parents=True, exist_ok=True)
        output_path = self._output_dir / filename

        df = self._build_dataframe(companies)
        df = self._compute_missing_ratios(df)
        stats = self._compute_summary_stats(df)

        initial_df: Optional[pd.DataFrame] = None
        if initial_company is not None:
            initial_df = self._build_dataframe([initial_company])
            initial_df = self._compute_missing_ratios(initial_df)

        self._write_excel(df, stats, output_path, initial_df=initial_df)

        logger.info(
            "Excel exporté : %s (%d sociétés)", output_path, len(df)
        )
        return output_path

    # ------------------------------------------------------------------
    # Construction du DataFrame
    # ------------------------------------------------------------------

    def _build_dataframe(self, companies: list[dict]) -> pd.DataFrame:
        """Construit un DataFrame ordonné à partir de la liste de sociétés.

        Sélectionne et renomme uniquement les colonnes définies dans _COLUMN_MAP.
        Les clés absentes sont remplacées par NaN.

        Args:
            companies: Liste de dictionnaires issus du Module 3.

        Returns:
            DataFrame avec les colonnes renommées selon _COLUMN_MAP.
        """
        rows = []
        for company in companies:
            row = {internal: company.get(internal) for internal in _COLUMN_MAP}
            rows.append(row)

        df = pd.DataFrame(rows)
        df.rename(columns=_COLUMN_MAP, inplace=True)

        # Conversion des types numériques (None → NaN)
        for col in _STAT_COLUMNS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        if "Exercice" in df.columns:
            df["Exercice"] = pd.to_numeric(df["Exercice"], errors="coerce")

        logger.debug("DataFrame construit : %d lignes, %d colonnes", *df.shape)
        return df

    def _compute_missing_ratios(self, df: pd.DataFrame) -> pd.DataFrame:
        """Dérive les ratios manquants si des colonnes sources sont disponibles.

        Règle appliquée : si marge_brute est NaN mais que les colonnes
        'Chiffre d'Affaires' et 'Résultat Brut' existent, le ratio est calculé.
        Cette méthode est un point d'extension — le pipeline actuel ne produit
        pas ces colonnes brutes, mais elles pourraient être ajoutées.

        Args:
            df: DataFrame avec les colonnes renommées.

        Returns:
            Le même DataFrame avec d'éventuels ratios comblés.
        """
        # Extension future : calcul depuis données brutes si disponibles
        # Ex : df["Marge Brute"].fillna(df["Résultat Brut"] / df["CA"], inplace=True)
        n_missing = df[_STAT_COLUMNS].isna().sum().sum()
        if n_missing:
            logger.debug("%d valeur(s) KPI manquante(s) après extraction LLM", n_missing)
        return df

    # ------------------------------------------------------------------
    # Statistiques de synthèse
    # ------------------------------------------------------------------

    def _compute_summary_stats(
        self, df: pd.DataFrame
    ) -> dict[str, dict[str, Optional[float]]]:
        """Calcule les statistiques descriptives du peer group pour les colonnes numériques.

        Args:
            df: DataFrame des entreprises.

        Returns:
            Dict {stat_label: {col_label: valeur}} pour chaque statistique.
        """
        stats: dict[str, dict[str, Optional[float]]] = {}
        numeric_df = df[[c for c in _STAT_COLUMNS if c in df.columns]]

        for label, method in _SUMMARY_LABELS:
            if method == "median":
                values = numeric_df.median(numeric_only=True)
            elif method == "mean":
                values = numeric_df.mean(numeric_only=True)
            elif method == "min":
                values = numeric_df.min(numeric_only=True)
            else:  # max
                values = numeric_df.max(numeric_only=True)

            stats[label] = {
                col: (float(values[col]) if col in values.index and pd.notna(values[col]) else None)
                for col in _STAT_COLUMNS
            }

        return stats

    # ------------------------------------------------------------------
    # Écriture Excel
    # ------------------------------------------------------------------

    def _write_excel(
        self,
        df: pd.DataFrame,
        stats: dict[str, dict[str, Optional[float]]],
        output_path: Path,
        initial_df: Optional[pd.DataFrame] = None,
    ) -> None:
        """Écrit le DataFrame et les statistiques dans un fichier Excel formaté.

        Args:
            df: DataFrame des entreprises avec les KPIs.
            stats: Statistiques de synthèse calculées par _compute_summary_stats.
            output_path: Chemin du fichier .xlsx de destination.
            initial_df: DataFrame (1 ligne) de la société de référence, ou None.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Benchmark KPIs"

        columns = list(df.columns)
        self._write_header_row(ws, columns)

        current_row = 2
        if initial_df is not None:
            self._write_initial_row(ws, initial_df, columns, current_row)
            current_row += 1
            self._write_peers_separator(ws, current_row, len(columns))
            current_row += 1

        last_data_row = self._write_data_rows(ws, df, columns, start_row=current_row)
        self._write_separator_row(ws, last_data_row + 1, len(columns))
        self._write_summary_rows(ws, stats, columns, start_row=last_data_row + 2)
        self._auto_fit_columns(ws, df, columns, initial_df=initial_df)

        # Gèle la première ligne (en-tête)
        ws.freeze_panes = "A2"

        try:
            wb.save(output_path)
        except OSError as exc:
            raise OSError(f"Impossible d'écrire le fichier Excel : {exc}") from exc

    def _write_header_row(self, ws, columns: list[str]) -> None:
        """Écrit et formate la ligne d'en-tête.

        Args:
            ws: Feuille openpyxl active.
            columns: Noms de colonnes dans l'ordre d'affichage.
        """
        ws.row_dimensions[1].height = 22
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER

    def _write_data_rows(
        self, ws, df: pd.DataFrame, columns: list[str], start_row: int = 2
    ) -> int:
        """Écrit les lignes de données avec alternance de couleurs.

        Args:
            ws: Feuille openpyxl active.
            df: DataFrame des entreprises.
            columns: Noms de colonnes dans l'ordre d'affichage.
            start_row: Première ligne disponible pour les données.

        Returns:
            Numéro de la dernière ligne écrite.
        """
        alt_fill = PatternFill(fill_type="solid", fgColor="EBF3FB")  # bleu très pâle

        for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row):
            ws.row_dimensions[row_idx].height = 18
            fill = alt_fill if row_idx % 2 == 0 else None

            for col_idx, col_name in enumerate(columns, start=1):
                value = row[col_name]
                # Convertit NaN en None pour qu'openpyxl laisse la cellule vide
                if pd.isna(value) if not isinstance(value, str) else False:
                    value = None

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = _DATA_FONT
                cell.border = _THIN_BORDER
                cell.alignment = _CENTER if col_name != "Société" else _LEFT

                if fill:
                    cell.fill = fill

                if col_name in _NUMERIC_FORMATS and value is not None:
                    cell.number_format = _NUMERIC_FORMATS[col_name]

        return start_row - 1 + len(df)

    def _write_initial_row(
        self, ws, initial_df: pd.DataFrame, columns: list[str], row: int
    ) -> None:
        """Écrit la ligne de la société de référence avec un style doré distinct.

        Args:
            ws: Feuille openpyxl active.
            initial_df: DataFrame (1 ligne) de la société initiale.
            columns: Noms de colonnes dans l'ordre d'affichage.
            row: Numéro de la ligne à écrire.
        """
        ws.row_dimensions[row].height = 18
        data_row = initial_df.iloc[0]

        for col_idx, col_name in enumerate(columns, start=1):
            value = data_row[col_name] if col_name in data_row.index else None
            if pd.isna(value) if not isinstance(value, str) else False:
                value = None

            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = _INITIAL_FONT
            cell.fill = _INITIAL_FILL
            cell.border = _THIN_BORDER
            cell.alignment = _LEFT if col_name == "Société" else _CENTER

            if col_name in _NUMERIC_FORMATS and value is not None:
                cell.number_format = _NUMERIC_FORMATS[col_name]

    def _write_peers_separator(self, ws, row: int, n_cols: int) -> None:
        """Insère une ligne de séparation visuelle entre la société initiale et les comparables.

        Args:
            ws: Feuille openpyxl active.
            row: Numéro de la ligne de séparation.
            n_cols: Nombre de colonnes du tableau.
        """
        ws.row_dimensions[row].height = 16
        label = "─── Groupe de comparables ───"
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(
                row=row,
                column=col_idx,
                value=label if col_idx == 1 else None,
            )
            cell.fill = _PEERS_SEPARATOR_FILL
            cell.font = _PEERS_SEPARATOR_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _LEFT if col_idx == 1 else _CENTER

    def _write_separator_row(self, ws, row: int, n_cols: int) -> None:
        """Insère une ligne de séparation visuelle avant les statistiques.

        Args:
            ws: Feuille openpyxl active.
            row: Numéro de la ligne de séparation.
            n_cols: Nombre de colonnes du tableau.
        """
        ws.row_dimensions[row].height = 16
        label = "─── Synthèse du peer group ───"
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(
                row=row,
                column=col_idx,
                value=label if col_idx == 1 else None,
            )
            cell.fill = _SEPARATOR_FILL
            cell.font = _SEPARATOR_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _LEFT if col_idx == 1 else _CENTER

    def _write_summary_rows(
        self,
        ws,
        stats: dict[str, dict[str, Optional[float]]],
        columns: list[str],
        start_row: int,
    ) -> None:
        """Écrit les lignes de statistiques (médiane, moyenne, min, max).

        Args:
            ws: Feuille openpyxl active.
            stats: {stat_label: {col_label: valeur}} calculé par _compute_summary_stats.
            columns: Noms de colonnes dans l'ordre d'affichage.
            start_row: Première ligne disponible pour les statistiques.
        """
        for row_offset, (label, _method) in enumerate(_SUMMARY_LABELS):
            row = start_row + row_offset
            ws.row_dimensions[row].height = 18
            col_values = stats.get(label, {})

            for col_idx, col_name in enumerate(columns, start=1):
                if col_idx == 1:
                    value = label
                    alignment = _LEFT
                elif col_name in col_values:
                    value = col_values[col_name]
                    alignment = _CENTER
                else:
                    value = None
                    alignment = _CENTER

                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = _SUMMARY_FONT
                cell.fill = _SUMMARY_FILL
                cell.border = _THIN_BORDER
                cell.alignment = alignment

                if col_name in _NUMERIC_FORMATS and value is not None:
                    cell.number_format = _NUMERIC_FORMATS[col_name]

    def _auto_fit_columns(
        self,
        ws,
        df: pd.DataFrame,
        columns: list[str],
        initial_df: Optional[pd.DataFrame] = None,
    ) -> None:
        """Ajuste la largeur de chaque colonne à son contenu le plus long.

        Prend en compte l'en-tête, la société initiale, les données, et les libellés de synthèse.

        Args:
            ws: Feuille openpyxl active.
            df: DataFrame des entreprises comparables.
            columns: Noms de colonnes dans l'ordre d'affichage.
            initial_df: DataFrame (1 ligne) de la société initiale, ou None.
        """
        summary_labels = [label for label, _ in _SUMMARY_LABELS]

        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_name))

            def _cell_str(val: object, col: str) -> str:
                if col == "Marge Brute":
                    return f"{val:.2%}"
                if col in ("CAPEX (M)", "Ratio Endettement"):
                    return f"{val:,.1f}"
                return str(val)

            for source_df in ([initial_df, df] if initial_df is not None else [df]):
                if source_df is not None and col_name in source_df.columns:
                    for val in source_df[col_name].dropna():
                        max_len = max(max_len, len(_cell_str(val, col_name)))

            if col_idx == 1:
                for lbl in summary_labels:
                    max_len = max(max_len, len(lbl))
                max_len = max(max_len, len("─── Synthèse du peer group ───"))
                max_len = max(max_len, len("─── Groupe de comparables ───"))

            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
